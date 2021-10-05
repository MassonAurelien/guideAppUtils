# coding=utf-8

#Partie Importation et Exportation Description BD et Excel
import sqlite3
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties,PageSetupProperties
from openpyxl.styles import Font
from database import connection,updateInsertTable
from pathlib import Path

chemin = input("Chemin d'accès BD NAMIP: ")
#chemin = "C:/Users/aurel/github/guideApp/assets/database/NAMIP.db"
con,cur = connection(chemin)

def initTableau():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Année'
    ws['B1'] = 'Nom'
    ws['C1'] = 'Type'
    ws['D1'] = 'DescFR'
    ws['E1'] = 'DescEN'
    ws['F1'] = 'DescNL'
    ws['G1'] = 'ID'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    ws['F1'].font = Font(bold=True)
    ws['G1'].font = Font(bold=True)
    return ws,wb

def ajouterData(ws):
    for row in cur.execute('SELECT Annee,Nom,TYPE,DescFR,DescEN,DescNL,ID FROM GENERAL;'):
        tab = []
        for x in row :
            tab.append(x)
        ws.append(tab)

def saveData(wb,chemin):
    wb.save(chemin+"/"+"namip.xlsx")

def recupData(ws):
    list = []
    for i in range (2,ws.max_row+1):
        listIntermediaire = []
        for j in range (1,ws.max_column+1):
            cellObj = ws.cell(row=i,column=j)
            listIntermediaire.append(cellObj.value)
        list.append(listIntermediaire)
    return list

action = input("A : BD vers excel ou B : Excel vers BD ?")
if(action == "B"):
    ws,wb = initTableau()
    ajouterData(ws)
    chemin = input("Donnez le chemin où enregistrer ?")
    if Path(chemin).exists() and Path(chemin).is_dir():
        saveData(wb,chemin)
        print("Ecriture dans un fichier Excel terminé")
    else :
        print("Le chemin donné n'existe pas")
elif(action == "A"):
    chemin = input("Entrez le chemin d'accès du fichier Excel:")
    if Path(chemin).exists() and Path(chemin).is_file() and chemin.find(".xlsx") !=-1:
        wb = load_workbook(chemin)
        ws = wb.active
        listDesc = recupData(ws)
        requete = "UPDATE GENERAL SET Annee = ?, Nom = ?, TYPE = ?, DescFR = ?, DescEN = ? ,DescNL = ? WHERE ID = ?;"
        updateInsertTable(con,cur,requete,listDesc)
        con.close()
        print("Insertion dans la BD terminée")
    else:
        print("Le chemin d'accès n'est pas bon")
else:
    print("Cette Action n'existe pas")